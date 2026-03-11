from pathlib import Path
import csv
import random
import re
import string
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook, load_workbook


base = Path(__file__).resolve().parents[1]
prompt_file = base / "prompts" / "_daily" / "Shipment_trackingNumber_prompt.txt"
completed_file = base / "output" / "Shipment-Tracking_completed.xlsx"


def parse_prompt_paths(path: Path):
    txt = path.read_text(encoding="utf-8")
    vals = {}
    for line in txt.splitlines():
        if "=" in line:
            key, value = line.split("=", 1)
            k = key.strip().upper()
            if k in {"FILE_NAME", "CHANNEL_ACCOUNT_MAPPING_FILE", "OUTPUT_FOLDER"}:
                vals[k] = value.strip()

    missing = {"FILE_NAME", "CHANNEL_ACCOUNT_MAPPING_FILE", "OUTPUT_FOLDER"} - set(vals)
    if missing:
        raise RuntimeError(f"Missing required prompt constants: {sorted(missing)}")
    return vals


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).lower()


def parse_channel_variants(raw_channel: str):
    raw = (raw_channel or "").strip()
    if not raw:
        return "", ""

    left = raw
    right = ""
    m = re.match(r"^(.*?)\s*\((.*?)\)\s*$", raw)
    if m:
        left = m.group(1).strip()
        right = m.group(2).strip()

    return left, right


def load_mapping(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        raise RuntimeError("Mapping file is empty")

    cols = list(rows[0].keys())
    lower = {str(c).strip().lower(): c for c in cols}

    num_col = (
        lower.get("channelaccountnum")
        or lower.get("channel account num")
        or lower.get("channel_account_num")
    )
    account_name_col = (
        lower.get("channelaccountname")
        or lower.get("channel account name")
        or lower.get("channel_account_name")
    )
    channel_name_col = (
        lower.get("channelname")
        or lower.get("channel name")
        or lower.get("channel_name")
    )

    if not num_col or not account_name_col:
        raise RuntimeError(f"Missing mapping columns. Found: {cols}")

    by_num = {}
    by_account_name = {}
    by_channel_name = {}

    for r in rows:
        num = str(r.get(num_col) or "").strip()
        account_name = str(r.get(account_name_col) or "").strip()
        channel_name = str(r.get(channel_name_col) or "").strip() if channel_name_col else ""

        if not num:
            continue

        by_num[num] = account_name or num

        if account_name:
            by_account_name[normalize_text(account_name)] = num
        if channel_name:
            by_channel_name[normalize_text(channel_name)] = num

    return by_num, by_account_name, by_channel_name


def index_headers(ws):
    headers = [str(c.value or "").strip() for c in ws[1]]
    return headers, {h: i for i, h in enumerate(headers)}


def find_col(index, *names):
    for n in names:
        if n in index:
            return index[n]
    return None


def ensure_col(ws, headers, index, name):
    if name in index:
        return index[name]

    headers.append(name)
    col_idx_1_based = len(headers)
    ws.cell(row=1, column=col_idx_1_based, value=name)
    index[name] = col_idx_1_based - 1
    return index[name]


def gen_ups(rng):
    chars = string.ascii_uppercase + string.digits
    return "1Z" + "".join(rng.choice(chars) for _ in range(16))


def gen_fedex(rng):
    return "".join(rng.choice(string.digits) for _ in range(rng.choice([12, 15])))


def gen_usps(rng):
    return "".join(rng.choice(string.digits) for _ in range(rng.choice([20, 21, 22])))


def gen_dhl(rng):
    return "".join(rng.choice(string.digits) for _ in range(10))


def gen_track(carrier, rng):
    c = (carrier or "").strip().upper()
    if c == "FEDEX":
        return gen_fedex(rng)
    if c == "USPS":
        return gen_usps(rng)
    if c == "DHL":
        return gen_dhl(rng)
    return gen_ups(rng)


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", (value or "").strip()).rstrip(".")
    return cleaned or "UNKNOWN_CHANNEL"


def get_cell_str(row_cells, idx):
    if idx is None:
        return ""
    return str(row_cells[idx].value or "").strip()


def resolve_channel_account_num(row_cells, channel_account_col, channel_col, by_account_name, by_channel_name):
    direct = get_cell_str(row_cells, channel_account_col)
    if direct:
        return direct

    channel_raw = get_cell_str(row_cells, channel_col)
    if not channel_raw:
        return ""

    left, right = parse_channel_variants(channel_raw)

    for candidate in [left, right, channel_raw]:
        key = normalize_text(candidate)
        if not key:
            continue
        if key in by_account_name:
            return by_account_name[key]
        if key in by_channel_name:
            return by_channel_name[key]

    return ""


def main():
    cfg = parse_prompt_paths(prompt_file)
    input_file = base / cfg["FILE_NAME"]
    mapping_file = base / cfg["CHANNEL_ACCOUNT_MAPPING_FILE"]
    output_folder = base / cfg["OUTPUT_FOLDER"]

    output_folder.mkdir(parents=True, exist_ok=True)
    for old in output_folder.glob("*.xlsx"):
        old.unlink()

    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file.as_posix()}")

    by_num, by_account_name, by_channel_name = load_mapping(mapping_file)

    wb = load_workbook(input_file)
    ws = wb.active

    headers, index = index_headers(ws)

    order_col = find_col(index, "channelOrderID", "Channel Order ID", "ChannelOrderID")
    if order_col is None:
        raise RuntimeError(f"Missing required order id column. Headers: {headers}")

    carrier_col = find_col(index, "Carrier")
    tracking_col = find_col(index, "Tracking Number")
    channel_account_col = find_col(index, "ChannelAccountNum")
    channel_col = find_col(index, "Channel", "channel", "Channel Name", "channelName")

    if tracking_col is None:
        tracking_col = ensure_col(ws, headers, index, "Tracking Number")

    if channel_account_col is None:
        channel_account_col = ensure_col(ws, headers, index, "ChannelAccountNum")

    rng = random.Random(20260310)

    rows = list(ws.iter_rows(min_row=2, max_col=len(headers)))
    order_groups = defaultdict(list)
    for row in rows:
        order_id = get_cell_str(row, order_col)
        order_groups[order_id].append(row)

    used_tracking = set()
    for row in rows:
        t = get_cell_str(row, tracking_col)
        if t:
            used_tracking.add(t)

    orders_new_tracking = 0
    orders_propagated = 0
    cells_filled = 0
    channel_account_derived = 0

    for order_id, group in order_groups.items():
        existing = []
        for row in group:
            t = get_cell_str(row, tracking_col)
            if t:
                existing.append(t)

        if existing:
            chosen = existing[0]
            if any(x != chosen for x in existing[1:]):
                raise RuntimeError(f"Conflicting existing tracking numbers within order id={order_id}")

            for row in group:
                if not get_cell_str(row, tracking_col):
                    row[tracking_col].value = chosen
                    cells_filled += 1
            orders_propagated += 1
        else:
            carrier = ""
            if carrier_col is not None:
                carrier = get_cell_str(group[0], carrier_col)

            t = gen_track(carrier, rng)
            while t in used_tracking:
                t = gen_track(carrier, rng)
            used_tracking.add(t)

            for row in group:
                row[tracking_col].value = t
                cells_filled += 1
            orders_new_tracking += 1

        for row in group:
            if not get_cell_str(row, channel_account_col):
                resolved = resolve_channel_account_num(
                    row,
                    channel_account_col,
                    channel_col,
                    by_account_name,
                    by_channel_name,
                )
                if resolved:
                    row[channel_account_col].value = resolved
                    channel_account_derived += 1

    completed_file_written = completed_file
    try:
        wb.save(completed_file_written)
    except PermissionError:
        completed_file_written = (
            base
            / "output"
            / f"Shipment-Tracking_completed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb.save(completed_file_written)

    out_headers = [h for i, h in enumerate(headers) if i != channel_account_col]
    by_channel_account = defaultdict(list)

    all_data_rows = [list(r) for r in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True)]
    for r in all_data_rows:
        account_num = str(r[channel_account_col] or "").strip()
        if not account_num:
            account_num = "UNKNOWN_CHANNEL"
        by_channel_account[account_num].append(r)

    written_files = []
    for account_num, channel_rows in sorted(by_channel_account.items()):
        channel_name = by_num.get(account_num, account_num)
        out_file = output_folder / f"{sanitize_filename(channel_name)}.xlsx"

        nwb = Workbook()
        nws = nwb.active
        nws.append(out_headers)
        for row in channel_rows:
            nws.append([v for i, v in enumerate(row) if i != channel_account_col])
        nwb.save(out_file)
        written_files.append(out_file.name)

    # Validation: each tracking number maps to only one order id.
    tracking_to_orders = defaultdict(set)
    order_to_tracking = defaultdict(set)
    for r in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        oid = str(r[order_col] or "").strip()
        tr = str(r[tracking_col] or "").strip()
        if tr:
            tracking_to_orders[tr].add(oid)
        if oid:
            order_to_tracking[oid].add(tr)

    tracking_reused_across_orders = sum(1 for orders in tracking_to_orders.values() if len(orders) > 1)
    order_tracking_violations = sum(1 for tracks in order_to_tracking.values() if len(tracks) > 1)

    print("prompt_file", prompt_file.as_posix())
    print("input_file", input_file.as_posix())
    print("mapping_file", mapping_file.as_posix())
    print("output_folder", output_folder.as_posix())
    print("completed_file", completed_file_written.as_posix())
    print("rows_total", len(all_data_rows))
    print("orders_total", len(order_groups))
    print("orders_with_new_tracking_generated", orders_new_tracking)
    print("orders_with_existing_tracking_propagated", orders_propagated)
    print("cells_filled", cells_filled)
    print("channel_accountnum_derived", channel_account_derived)
    print("tracking_unique_values", len(tracking_to_orders))
    print("tracking_reused_across_orders", tracking_reused_across_orders)
    print("order_tracking_violations", order_tracking_violations)
    print("split_files_written", len(written_files))
    print("split_files", sorted(written_files))


if __name__ == "__main__":
    main()
