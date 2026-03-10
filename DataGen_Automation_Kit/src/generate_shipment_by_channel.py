from __future__ import annotations

import csv
import random
import re
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook, load_workbook

BASE = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE / "output"
SHIPMENT_DIR = OUTPUT_DIR / "Shipment"
PROMPT_FILE = BASE / "prompts" / "_daily" / "salesOrder_daily_w_all_customerType_prompt.md"

OUTPUT_HEADERS = [
    "Channel Order ID",
    "Ship Date",
    "TimeZone",
    "Carrier",
    "Tracking Number",
    "Shipping Service",
    "2nd Tracking Number",
    "Package",
    "Shipping Fee",
    "Weight",
    "Length",
    "Width",
    "Height",
    "Note",
    "SKU",
    "Ship Qty",
]


def parse_prompt_paths(prompt_path: Path) -> tuple[Path, Path]:
    text = prompt_path.read_text(encoding="utf-8")

    source_match = re.search(r"Source file:\s*\n\s*-\s*([^\r\n]+)", text, flags=re.IGNORECASE)
    output_match = re.search(r"Output folder:\s*([^\r\n]+)", text, flags=re.IGNORECASE)

    source_rel = source_match.group(1).strip() if source_match else "data/Shipment_From_SO.xlsx"
    output_rel = output_match.group(1).strip() if output_match else "output/Shipment/"

    return BASE / source_rel, BASE / output_rel


def read_rows_from_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(min_row=1, values_only=True)
    headers = [str(cell or "").strip() for cell in next(iterator)]

    rows: list[dict[str, str]] = []
    for values in iterator:
        rows.append({headers[index]: str(value or "").strip() for index, value in enumerate(values)})
    return rows


def read_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".xlsx":
        return read_rows_from_xlsx(path)

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        return [{(k or "").strip(): (v or "").strip() for k, v in row.items()} for row in reader]


def pick_value(row: dict[str, str], keys: list[str]) -> str:
    for key in keys:
        value = row.get(key, "").strip()
        if value:
            return value
    return ""


def resolve_channel(row: dict[str, str]) -> str:
    channel = pick_value(row, ["Channel", "channel", "Channel Name", "channelName"])
    if channel:
        return channel

    channel_num = row.get("ChannelNum", "").strip()
    if channel_num:
        return channel_num

    channel_account_num = row.get("ChannelAccountNum", "").strip()
    if channel_account_num:
        return channel_account_num

    return "UNKNOWN_CHANNEL"


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", value).strip().rstrip(".")
    return cleaned or "UNKNOWN_CHANNEL"


def generate_unique_fedex_tracking(existing: set[str], rng: random.Random) -> str:
    while True:
        length = rng.choice([12, 15])
        tracking = "".join(str(rng.randint(0, 9)) for _ in range(length))
        if tracking not in existing:
            existing.add(tracking)
            return tracking


def build_output_row(source_row: dict[str, str], tracking_number: str) -> dict[str, str]:
    ship_date = pick_value(source_row, ["Est.Ship Date", "ShipDate", "Ship Date"])
    return {
        "Channel Order ID": pick_value(source_row, ["ChannelOrderID", "Channel Order ID"]),
        "Ship Date": ship_date,
        "TimeZone": "UTC-8",
        "Carrier": "FedEx",
        "Tracking Number": tracking_number,
        "Shipping Service": "",
        "2nd Tracking Number": "",
        "Package": "",
        "Shipping Fee": "",
        "Weight": "",
        "Length": "",
        "Width": "",
        "Height": "",
        "Note": "",
        "SKU": pick_value(source_row, ["SKU"]),
        "Ship Qty": pick_value(source_row, ["OrderQty", "Order Qty"]),
    }


def write_channel_file(channel: str, rows: list[dict[str, str]]) -> Path:
    SHIPMENT_DIR.mkdir(parents=True, exist_ok=True)
    file_name = f"Shipment_{sanitize_filename(channel)}.xlsx"
    output_file = SHIPMENT_DIR / file_name

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(OUTPUT_HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in OUTPUT_HEADERS])
    workbook.save(output_file)

    return output_file


def main() -> None:
    source_file, output_dir = parse_prompt_paths(PROMPT_FILE)
    global SHIPMENT_DIR
    SHIPMENT_DIR = output_dir

    if not source_file.exists():
        raise FileNotFoundError(f"Source file not found: {source_file.as_posix()}")

    source_rows = read_rows(source_file)

    sorted_rows = sorted(
        source_rows,
        key=lambda row: (
            resolve_channel(row),
            pick_value(row, ["ChannelOrderID", "Channel Order ID"]),
            pick_value(row, ["SKU"]),
        ),
    )

    rng = random.Random(20260309)
    used_tracking_numbers: set[str] = set()

    grouped_output_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in sorted_rows:
        channel = resolve_channel(row)
        tracking_number = generate_unique_fedex_tracking(used_tracking_numbers, rng)
        grouped_output_rows[channel].append(build_output_row(row, tracking_number))

    for old_file in SHIPMENT_DIR.glob("Shipment_*.csv"):
        old_file.unlink()
    for old_file in SHIPMENT_DIR.glob("Shipment_*.xlsx"):
        old_file.unlink()

    written_files = []
    for channel in sorted(grouped_output_rows.keys()):
        output_file = write_channel_file(channel, grouped_output_rows[channel])
        written_files.append(output_file.name)

    print("source_file", source_file.as_posix())
    print("source_rows", len(source_rows))
    print("channels", len(grouped_output_rows))
    print("written_files", len(written_files))
    print("tracking_numbers_unique", len(used_tracking_numbers))
    print("shipment_files", written_files)


if __name__ == "__main__":
    main()
