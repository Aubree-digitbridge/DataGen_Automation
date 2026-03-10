from __future__ import annotations
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl

@dataclass
class TabularTemplate:
    path: Path
    headers: List[str]
    sheet: Optional[str] = None

def read_csv_headers(path: str) -> TabularTemplate:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Template not found: {p}")

    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, None)

    if not headers:
        raise ValueError(f"No header row found in: {p}")

    headers = [h.strip() for h in headers]
    if any(h == "" for h in headers):
        raise ValueError(f"Blank header detected in: {p}")

    return TabularTemplate(path=p, headers=headers)

def read_xlsx_headers(path: str, sheet_name: str | None = None) -> TabularTemplate:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Template not found: {p}")

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # pick first sheet
        ws = wb[wb.sheetnames[0]]

    # assume headers are in row 1
    headers = []
    for cell in ws[1]:
        val = (cell.value or "")
        headers.append(str(val).strip())

    if not any(headers):
        raise ValueError(f"No header row found in xlsx: {p} (sheet={ws.title})")

    headers = [h for h in headers if h != ""]
    return TabularTemplate(path=p, headers=headers, sheet=ws.title)

def read_headers(path: str, sheet_name: str | None = None) -> TabularTemplate:
    p = Path(path)
    if p.suffix.lower() in [".csv", ".txt"]:
        return read_csv_headers(path)
    if p.suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        return read_xlsx_headers(path, sheet_name=sheet_name)
    raise ValueError(f"Unsupported template type: {p.suffix} for {p}")
