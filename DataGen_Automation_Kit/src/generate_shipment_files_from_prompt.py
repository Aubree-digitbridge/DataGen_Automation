from __future__ import annotations

import random
import re
import string
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parents[1]
INPUT_FILE = BASE / "data" / "Shipment_From_SO.xlsx"
OUTPUT_DIR = BASE / "output"
CHANNEL_OUTPUT_DIR = OUTPUT_DIR / "Shipment"
UPDATED_MASTER_FILE = OUTPUT_DIR / "Shipment_From_SO_updated.xlsx"
ZIP_FILE = OUTPUT_DIR / "Shipment_Tracking_Output.zip"

CHANNEL_HEADERS = [
    "Channel Order ID",
    "Ship Date",
    "TimeZone",
    "Carrier",
    "Tracking Number",
    "Shipping Service",
    "2nd Tracking Number",
    "Package",
    "Shipping Fee",
    "Weight",
    "Length",
    "Width",
    "Height",
    "Note",
    "SKU",
    "Ship Qty",
]


def sanitize_channel_filename(channel_value: str) -> str:
    # Keep prompt-compatible naming: replace spaces and invalid chars with underscores.
    value = (channel_value or "").strip()
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r'[<>:"/\\|?*]', "_", value)
    value = re.sub(r"_+", "_", value).strip("_.")
    return value or "UNKNOWN_CHANNEL"


def gen_ups(rng: random.Random) -> str:
    chars = string.ascii_uppercase + string.digits
    return "1Z" + "".join(rng.choice(chars) for _ in range(16))


def gen_fedex(rng: random.Random) -> str:
    length = rng.choice([12, 15])
    return "".join(rng.choice(string.digits) for _ in range(length))


def gen_tracking_for_carrier(carrier: str, rng: random.Random) -> str:
    if carrier == "FedEx":
        return gen_fedex(rng)
    return gen_ups(rng)


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Missing source file: {INPUT_FILE.as_posix()}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CHANNEL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    headers = [str(c.value or "").strip() for c in ws[1]]
    index = {h: i for i, h in enumerate(headers)}

    order_col = index.get("Channel Order ID")
    channel_col = index.get("Channel")
    ship_date_col = index.get("Est.Ship Date")
    order_qty_col = index.get("Order Qty")
    sku_col = index.get("SKU")

    if order_col is None or channel_col is None or ship_date_col is None or order_qty_col is None or sku_col is None:
        raise RuntimeError("Source is missing one or more required fields: Channel Order ID, Channel, Est.Ship Date, Order Qty, SKU")

    tracking_col = index.get("Tracking Number")
    if tracking_col is None:
        headers.append("Tracking Number")
        tracking_col = len(headers) - 1
        ws.cell(row=1, column=tracking_col + 1, value="Tracking Number")

    rows = list(ws.iter_rows(min_row=2, max_col=len(headers)))

    # Preserve original row order while grouping by order id.
    order_rows: dict[str, list] = defaultdict(list)
    for row in rows:
        order_id = str(row[order_col].value or "").strip()
        order_rows[order_id].append(row)

    rng = random.Random(20260310)
    used_tracking: set[str] = set()

    existing_by_order: dict[str, str] = {}
    for order_id, group in order_rows.items():
        values = []
        for row in group:
            v = str(row[tracking_col].value or "").strip()
            if v:
                values.append(v)
        if values:
            chosen = values[0]
            if any(v != chosen for v in values[1:]):
                raise RuntimeError(f"Conflicting existing tracking numbers in order group: {order_id}")
            existing_by_order[order_id] = chosen
            used_tracking.add(chosen)

    # Random carrier assignment per order group (UPS/FedEx), used for generating tracking format.
    carrier_by_order: dict[str, str] = {}
    for order_id in order_rows.keys():
        carrier_by_order[order_id] = rng.choice(["UPS", "FedEx"])

    tracking_generated = 0
    tracking_reused = 0

    for order_id, group in order_rows.items():
        if order_id in existing_by_order:
            tracking = existing_by_order[order_id]
            tracking_reused += 1
        else:
            carrier = carrier_by_order[order_id]
            tracking = gen_tracking_for_carrier(carrier, rng)
            while tracking in used_tracking:
                tracking = gen_tracking_for_carrier(carrier, rng)
            used_tracking.add(tracking)
            tracking_generated += 1

        for row in group:
            current = str(row[tracking_col].value or "").strip()
            if not current:
                row[tracking_col].value = tracking

    wb.save(UPDATED_MASTER_FILE)

    # Build channel output rows with exact schema and grouped order consistency.
    by_channel: dict[str, list[list]] = defaultdict(list)
    for row in rows:
        order_id = str(row[order_col].value or "").strip()
        channel = str(row[channel_col].value or "").strip()
        ship_date = str(row[ship_date_col].value or "").strip()
        sku = str(row[sku_col].value or "").strip()
        order_qty = str(row[order_qty_col].value or "").strip()
        tracking = str(row[tracking_col].value or "").strip()
        carrier = carrier_by_order.get(order_id, "UPS")

        out_row = [
            order_id,
            ship_date,
            "UTC-8",
            carrier,
            tracking,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            sku,
            order_qty,
        ]
        by_channel[channel].append(out_row)

    channel_files: list[Path] = []
    for old in CHANNEL_OUTPUT_DIR.glob("*.xlsx"):
        old.unlink()

    for channel, channel_rows in sorted(by_channel.items(), key=lambda kv: kv[0].lower()):
        out_name = f"{sanitize_channel_filename(channel)}.xlsx"
        out_file = CHANNEL_OUTPUT_DIR / out_name

        nwb = Workbook()
        nws = nwb.active
        nws.append(CHANNEL_HEADERS)
        for r in channel_rows:
            nws.append(r)
        nwb.save(out_file)
        channel_files.append(out_file)

    if ZIP_FILE.exists():
        ZIP_FILE.unlink()

    with zipfile.ZipFile(ZIP_FILE, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(UPDATED_MASTER_FILE, UPDATED_MASTER_FILE.name)
        for cf in channel_files:
            zf.write(cf, cf.name)

    print("updated_master", UPDATED_MASTER_FILE.as_posix())
    print("zip_file", ZIP_FILE.as_posix())
    print("rows_processed", len(rows))
    print("tracking_generated", tracking_generated)
    print("tracking_reused", tracking_reused)
    print("channel_files_produced", len(channel_files))
    print("channel_file_names", [p.name for p in channel_files])


if __name__ == "__main__":
    main()
